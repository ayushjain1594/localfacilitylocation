import pandas as pd
from openpyxl import load_workbook
import os
from math import radians, cos, sin, acos, asin

from routeflows import RouteFlows
from facilitylocation import FacilityLocationModel

class CapEx:
	def __init__(self, filename='InputData - Copy.xlsx'):
		
		self.readData(filename)
		self.processData()

	def readData(self, filename):
		'''Method reads following input tables from excel file
		named 'InputData.xlsx' with corresponding sheet names
		'''
		try:
			self.df = {}
			workbook = pd.ExcelFile(filename)
			self.df['customers'] = pd.read_excel(
				workbook, 'Customers')
			self.df['sites'] = pd.read_excel(
				workbook, 'Sites')
			self.df['periods'] = pd.read_excel(
				workbook, 'Periods')
			self.df['demand'] = pd.read_excel(
				workbook, 'CustomerDemand')
			self.df['sitecapacity'] = pd.read_excel(
				workbook, 'SiteCapacity')
			self.df['scenarios'] = pd.read_excel(
				workbook, 'Scenarios')
		except IOError:
			print("Error occured reading data. Exiting")
			import sys
			sys.exit()


	def getListFromDataframe(self, key, column):
		return self.df[key][self.df[key].Status == \
			'Include'][column].tolist()


	def getDictFromDataframe(self, key, columns, df_override=None):
		if isinstance(df_override, pd.DataFrame):
			df_temp = df_override
		else:
			df_temp = self.df[key]
		if len(columns) == 2:
			return dict(df_temp[df_temp.Status == \
				'Include'][columns].values)
		elif len(columns) == 3:
			returndict = {}
			uniqueval = set(self.getListFromDataframe(
				key, columns[0])
			)
			for val in uniqueval:
				returndict[val] = self.getDictFromDataframe(
					key, columns[1:], 
					df_temp[df_temp[columns[0]] == val])
			return returndict


	def processData(self):
		'''Method reads information from dataframe and 
		converts into nicer data structure for ease of 
		use'''
		# location IDs
		self.customerID = self.getListFromDataframe(
			'customers', 'CustomerID')
		self.siteID = self.getListFromDataframe(
			'sites', 'SiteID')
		self.periodID = self.getListFromDataframe(
			'periods', 'PeriodID')

		# coordinates by location id
		self.customerLat = self.getDictFromDataframe(
			'customers', ['CustomerID', 'CustomerLatitude'])
		self.customerLon = self.getDictFromDataframe(
			'customers', ['CustomerID', 'CustomerLongitude'])
		self.siteLat = self.getDictFromDataframe(
			'sites', ['SiteID', 'SiteLatitude'])
		self.siteLon = self.getDictFromDataframe(
			'sites', ['SiteID', 'SiteLongitude'])

		# site capacity in terms of max shipments per day
		self.siteCapByPeriod = self.getDictFromDataframe(
			'sitecapacity', ['PeriodID', 'SiteID', 'Capacity'])

		# slack capacity on top of exiting capacity
		self.siteSlackCapByPeriod = self.getDictFromDataframe(
			'sitecapacity', ['PeriodID', 'SiteID', 'CapacitySlack'])

		# customer demand, these are average shipments per day by cust id
		self.customerDemByPeriod = self.getDictFromDataframe(
			'demand', ['PeriodID', 'CustomerID', 'Demand'])

		self.scenarioID = self.getListFromDataframe(
			'scenarios', 'ScenarioID')

		self.computeServiceDistances()


	def computeServiceDistances(self):
		'''Method computes distance between
		facilities and customers using great
		circle distance formula'''

		self.serviceDist = {}
		for sid in self.siteID:
			lat1 = radians(self.siteLat[sid])
			lon1 = radians(self.siteLon[sid])
			for cid in self.customerID:
				lat2 = radians(self.customerLat[cid])
				lon2 = radians(self.customerLon[cid])

				# radius of earth in miles - 3958.75
				self.serviceDist[(sid, cid)] = round(3958.75 * (
					acos(sin(lat1)*sin(lat2) \
						+ cos(lat1)*cos(lat2)*cos(lon1-lon2))), 
					2)


	def createFlows(self):
		flm = FacilityLocationModel(
			{
				'periodid': self.periodID,
				'siteid': self.siteID,
				'customerid': self.customerID,
				'sitecapbyperiod': self.siteCapByPeriod,
				'siteslackcapbyperiod': self.siteSlackCapByPeriod,
				'customerdembyperiod': self.customerDemByPeriod,
				'servicedist': self.serviceDist,
				'maxopensites': 4
			}
		)
		flm.modelProblem()
		flm.setParameters()
		flm.solveModel()
		flows, objectiveval = flm.extractSolution()
		return flows, objectiveval


	def solve(self):

		flows, flowcost = self.createFlows()

		flowrows = []
		pathrows = []
		clusters = []
		routes = []
		routepaths = []

		scid = self.scenarioID[0]

		for key, flow in flows.items():
			if flow > 0:
				pid, sid, cid = key

				# flows
				flowrows.append(
					[scid, pid, sid, cid, flow, 
					self.serviceDist[sid, cid], 
					flowcost.get(key)]
				)

				# paths
				pathid = "_".join([str(id_) for id_ in key])

				pathrows.append(
					[scid, pathid, pid, 'Site', 
					sid, self.siteLat[sid], 
					self.siteLon[sid], flow]
				)
				pathrows.append(
					[scid, pathid, pid, 'Customer', 
					cid, self.customerLat[cid], 
					self.customerLon[cid], flow]
				)
		df_flow = CapEx.putInDataFrame(
			flowrows, datafor='flows')
		df_path = CapEx.putInDataFrame(
			pathrows, datafor='flowpaths')

		routeflows = RouteFlows(
			self.siteLat, self.siteLon,
			self.customerLat, self.customerLon,
			self.siteID, self.customerID,
			7, 700
		)
		for pid in self.periodID:
			cluster_pid, route_pid, routepaths_pid = \
				routeflows.createFlowRoutes(
					df_flow[df_flow.PeriodID == pid],
					pid, scid
				)

			clusters.extend(cluster_pid)
			routes.extend(route_pid)
			routepaths.extend(routepaths_pid)
			break

		df_clusters = CapEx.putInDataFrame(
			clusters, datafor='clusters')
		df_routes = CapEx.putInDataFrame(
			routes, datafor='routes')
		df_routepaths = CapEx.putInDataFrame(
			routepaths, datafor='routepaths')

		addUpdateSheet(scid, 'OutputClusters.xlsx',
			{'Clusters': df_clusters})
		addUpdateSheet(scid, 'OutputRoutes.xlsx',
			{'Routes': df_routes})
		addUpdateSheet(scid, 'OutputRoutePaths.xlsx',
			{'RoutePaths': df_routepaths})
		addUpdateSheet(scid, 'OutputData.xlsx',
			{'OutputFlow': df_flow, 'OutputPath': df_path})


	@staticmethod
	def putInDataFrame(rows, datafor='flows'):
		"""
		Method to accept rows and create 
		pandas dataframe with columns based
		on type of data. Removes messy list
		of strings from core methods.
		args:
			rows: List of lists
				Each list within represents
				a row for require dataframe
		kwargs:
			datafor: str
				Specify what data rows are for.
				Must be among following list.
		"""
		columns_ = {
			'flows': ['ScenarioID', 'PeriodID', 'SiteID', 
				'CustomerID', 'FlowUnits', 'Distance', 
				'ObjectiveValue'],

			'flowpaths': ['ScenarioID', 'PathID', 'PeriodID', 
				'LocationType', 'LocationID', 'Latitude', 
				'Longitude', 'FlowUnits'],

			'clusters': ['ScenarioID', 'PeriodID', 
				'SiteID', 'ClusterID', 'CustomerID', 
				'Count', 'Weight Arcs', 'Weight Nodes'],

			'routes': ['ScenarioID', 'PeriodID', 
				'SiteID', 'RouteID', 'StopNumber', 
				'StopType', 'StopID', 'Distance', 
				'Cumulated Distance', 'LegType'],

			'routepaths': ['ScenarioID', 'PeriodID', 
				'SiteID', 'RouteID', 'RouteKey',
				'Latitude', 'Longitude', 'Distance']
		}
		if datafor in columns_:
			try:
				return pd.DataFrame(rows, 
					columns=columns_.get(datafor))
			except Exception as e:
				print("Something went wrong "+\
					f"creating dataframe for {datafor}"
				)
			return pd.DataFrame([], 
				columns=columns_.get(datafor))
		else:
			raise KeyError("Unexpected data received")
			return None


def addUpdateSheet(scid, file, sheetdfdict={}):
	for sheet, df in sheetdfdict.items():
		df.to_csv("test/"+str(scid)+str(file[:-5])+sheet+".csv", index=False)
	"""
	print(f"\nCalled to update {file}")
	try:
		wb = load_workbook(file)
		print('Load Workbook Done .. ')
	except IOError:
		return
	writer = pd.ExcelWriter(file, 
		engine='openpyxl')
	print('Writer Object Created')
	for sheet, df in sheetdfdict.items():
		print(f'Updating sheet {sheet}')
		if sheet not in wb.sheetnames:
			df_sheet = df
		else:
			data = wb[sheet].values
			df_sheet = pd.DataFrame(data, columns=next(data))
			#df_sheet = pd.read_excel(
			#	pd.ExcelFile(file),
			#	sheet)
			print("Finished reading data")
			df_sheet = df_sheet[~(df_sheet.ScenarioID == scid)]
			df_sheet = df_sheet.append(df, ignore_index=True)
			print("Finished writing data")
			wb.remove(wb[sheet])
		writer.book = wb
		df_sheet.to_excel(writer, sheet_name=sheet,
			index=False)
	
	writer.save()
	writer.close()
	print("Finished export")
	"""



def test():
	fcp = CapEx()
	fcp.solve()


if __name__ == '__main__':
	test()